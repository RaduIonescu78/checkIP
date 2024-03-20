import ipaddress
from ipaddress import ip_address, IPv6Address
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter.simpledialog import askstring
import save_attachments
import openpyxl
import xlrd
from openpyxl.utils import get_column_letter
import os

# Call the function to save the "All IP address.xls" file from email
folder_obj = save_attachments.SaveAtt("Reports")
save_all_att = folder_obj.save_attachments()

LARGE_FONT= ("Verdana", 12)
NORM_FONT= ("Verdana", 10)
SMALL_FONT= ("Verdana", 8)


# FUNCTIONS

# Function to create a pop-up message

def popupmsg(msg: object) -> object:
    popup = tk.Tk()
    popup.wm_title("!")
    label = ttk.Label(popup, text=msg, font=LARGE_FONT)
    label.pack(side="top", fill="x", pady=10)
    B1 = ttk.Button(popup, text="Okay", command = popup.destroy)
    B1.pack()
    popup.mainloop()

# Function to remove the zeros from the left
def remove_zeros(nr):
    # traverse the entire string
    for i in range(len(str(nr))):
        # check for the first non-zero character
        if nr[i] != '0' and nr != '0000':
            # return the remaining string
            res = nr[i::]
            return res
        # if the string is '0000' return '0'
        elif nr == '0000':
            res = '0'
            return res


# Function to get the column values for a column name
def get_columns_from_worksheet(sheet):
    return {
        cell.value: {
            'letter': get_column_letter(cell.column),
            'number': cell.column - 1
        } for cell in sheet[1] if cell.value
    }


# Function to check the duplicate IPs
dup_dict = {}
no_dup_list = []
def check_duplicate(ip):
    #dup_dict = {}
    global dup_dict
    #no_dup_list = []
    global no_dup_list
    dup = 0
    wb = xlrd.open_workbook((os.path.abspath(__file__)[0:len(os.path.abspath(__file__)) - 11]) + "IP_ADDRESSES.xls")
    sheet = wb.sheet_by_index(0)
    for i in range(sheet.nrows):
        for j in range(sheet.ncols):
            if sheet.cell_value(i, j) == ip:  # and sheet.cell_value(i, j - 1) != site_name:
                #print(sheet.cell_value(i, j) + " found on live network site:  " + sheet.cell_value(i, j - 1))
                dup_dict[sheet.cell_value(i, j)] = sheet.cell_value(i, j - 1)
                #popupmsg('The IP ' + sheet.cell_value(i, j) + " is already used in live network on site:  " + sheet.cell_value(i, j - 1))
                dup += 1
                #print("The IPs are already used in live on the sites:", dup_dict)
    if dup == 0 and ip is not None:
        no_dup_list.append(ip)
        #print("No duplicate IP found for:", no_dup_list)

# print("No duplicate IPP found for:", no_dup_list)
# print("The IPPs are already used in live on the sites:", dup_dict )

        #popupmsg('No duplicate IP found for : ' + ip)


# Function to change the IPs from FPT format to Huawei format

def change_ipv6_huawei(ad):
    ip_list = ad.split(":")  # Split the IP address into a list by ':'
    for i in range(0, len(ip_list)):  # check for the elements of the ip_list which starts with a '0'
        if ip_list[i].startswith("0"):
            ip_list[i] = remove_zeros(
                ip_list[i])
    # join the elements of the resulting ip_list into a string adding ':' between elements
    ip_huawei = ':'.join(str(e) for e in ip_list)  # IP address in Huawei format
    return ip_huawei


# Function to check if an IP is IPv4 of IPv6

def validipaddress(IP: str) -> str:
    try:
        return "IPv6" if type(ip_address(IP)) is ipaddress.IPv6Address else "IPv4"
    except ValueError:
        return "Invalid"


# Function Gets the last 3 characters from a list of strings

def get_rear(sub):
    return sub[-3:]

site_name = askstring("Insert Site name", "Site Name ").upper()
print(site_name)
#site_name = input("Please input the site name: ").upper()
# print("abspath", os.path.abspath(__file__))
popupmsg("Please select the Cramer file from COM5 in the next window")
file_path = filedialog.askopenfilename()  # pop-up window to open the Cramer xlsx file
wb = openpyxl.load_workbook(file_path, data_only=True)  # open the Cramer file
sheet = wb['Sheet0']

# Get the corresponding row in the sheet for the chosen site

row_to_check = ''
for row in sheet.values:
    if str(site_name[0] + 'XL' + get_rear(site_name)) in row or str(site_name[0] + 'XB' + get_rear(site_name)) in row or str(site_name[0] + 'AB' + get_rear(site_name)) in row or str(
            site_name[0] + 'XV' + get_rear(site_name)) in row or str(site_name[0] + 'AL' + get_rear(site_name)) in row or str(site_name[0] + 'BL' + get_rear(site_name)) in row:
        row_to_check += str(row)

print("Rows to be checked: ", row_to_check)

columns = get_columns_from_worksheet(sheet)  # variable to store the columns values
# print("Columns: ", COLUMNS)
ip_to_check = []  # list to store IPs to be checked

for cell in sheet[columns['OAM_IP']['letter']]:  # get the OAM_IP for the site to be checked
    if cell.value is not None and cell.value in row_to_check and cell.value not in ip_to_check:
        ip_to_check.append(cell.value)

#print("IP to check after 1st for:", ip_to_check)

for cell in sheet[columns['CUP_IP']['letter']]:  # get the CUP_IP for the site to be checked
    if cell.value is not None and cell.value in row_to_check and cell.value not in ip_to_check:
        # if cell.value in row_to_check:
        ip_to_check.append(cell.value)

# print("IP to check after 2nd for:", ip_to_check)

for cell in sheet[columns['CUP_INNER_IP']['letter']]:  # get the CUP_INNER_IP for the site to be checked
    if cell.value is not None and cell.value in row_to_check and cell.value not in ip_to_check:
        ip_to_check.append(cell.value)

#print("IPs to check OAM / CUP / CUP_INNER :", ip_to_check)
# popupmsg('IPs to check OAM / CUP / CUP_INNER :' + str(ip_to_check[0:]))
for ipaddr in ip_to_check:  # call the function to check duplicate IPs
    if validipaddress(ipaddr) == "IPv6":  # if IP is a IPv6
        check_duplicate(change_ipv6_huawei(ipaddr))  # change the ip from 2a01:08f1:e0a5:0075:0000:0000:0000:0002
        # to export from huawei U2020 format: 2a01:8f1:e0a5:75:0:0:0:2 and check for duplicate in live
    else:
        check_duplicate(ipaddr)  # if IPv4 check for duplicate in live
popupmsg('The next IPs are already used on site: ' + str(dup_dict)
+ '\n\n\nNo duplicate IP found for:' + str(no_dup_list))

# print("No duplicate IP found for:", no_dup_list)
# print("The IPs are already used in live on the sites:", dup_dict )
ip = '2a01:08f1:e4a5:042a:0000:0000:0000:0002'  # IP address
addr = ipaddress.ip_address(ip)  # IP address stripped (Ericsson format in ENM)

# print("IP address Ericsson ENM format:", addr)
